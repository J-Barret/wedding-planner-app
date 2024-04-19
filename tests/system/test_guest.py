from models.guest import GuestModel
from tests.base_test import BaseTest
from werkzeug.datastructures import Authorization
import json
import tempfile
import os
from openpyxl import Workbook
from io import BytesIO

def fill_excel(excel_path, expected_columns, data):
	wb = Workbook()  # create a workbook and select the active worksheet
	ws = wb.active
	for col, header in enumerate(expected_columns, start=1):
		ws.cell(row=1, column=col, value=header)
	for row, item in enumerate(data, start=2):
		for col, value in enumerate(item, start=1):
			ws.cell(row=row, column=col, value=value)
	wb.save(excel_path)
	wb.close()

class GuestTest(BaseTest):

	def setUp(self):
		super().setUp() #calling parent class setUp to avoid overwriting it with the instructions below
		fd1, self.excel_path = tempfile.mkstemp(suffix=".xlsx") #create a temporary Excel file
		expected_columns = ["name", "number", "email"]
		data = [
			("name1", "number1", "email1"),
			("name2", "number2", "email2"),
		]
		fill_excel(self.excel_path, expected_columns, data)

		fd2, self.odt_path = tempfile.mkstemp(suffix=".ods")  # create a temporary .ods file
		fd3, self.excel_empty_path = tempfile.mkstemp(suffix=".xlsx") #Excel blank file
		fd4, self.excel_wrong_format_path = tempfile.mkstemp(suffix=".xlsx") #Excel file with wrong columns (in this case, blank)
		expected_columns = ["name", "number"]
		data = [
			("name1", "number1"),
			("name2", "number2"),
		]
		fill_excel(self.excel_wrong_format_path, expected_columns, data)

		os.close(fd1) #important to do this to release the files (avoid issues when deleting them in tearDown()
		os.close(fd2)
		os.close(fd3)
		os.close(fd4)

		self.assertTrue(os.path.exists(self.excel_path))
		self.assertTrue(os.path.exists(self.odt_path))
		self.assertTrue(os.path.exists(self.excel_empty_path))
		self.assertTrue(os.path.exists(self.excel_wrong_format_path))

		#obtaining auth necessary to upload files
		with self.client as client:
			with self.app_context():
				client.post("/register", json={"name": "name1", "password": "password1"})
				request = client.post("/login", json={"name": "name1", "password": "password1"})
				access_token = json.loads(request.data)["access_token"]
				self.auth = Authorization(auth_type="bearer", token=access_token)

	def test_upload_ok(self):
		with self.client as client:
			with self.app_context():
				with open(self.excel_path, "rb") as file:
					filename = os.path.basename(self.excel_path)
					request = client.post("/upload", auth=self.auth, data={"file": (file, filename)}, content_type="multipart/form-data")
				self.assertEqual(request.status_code, 201)
				self.assertEqual(json.loads(request.data)["message"], "Guests added to database succesfully.")
				guest1 = GuestModel("name1", "number1", "email1", "1")
				guest2 = GuestModel("name2", "number2", "email2", "1")
				self.assertIsNotNone(GuestModel.find_by_name(guest1.name))
				self.assertIsNotNone(GuestModel.find_by_name(guest2.name))
	def test_file_missing(self):
		with self.client as client:
			with self.app_context():
				request = client.post("/upload", auth=self.auth)
				self.assertEqual(request.status_code, 400)
				self.assertEqual(json.loads(request.data)["message"], "No file part in the request")
	def test_file_extension_wrong(self):
		with self.client as client:
			with self.app_context():
				with open(self.odt_path, "rb") as file:
					filename = os.path.basename(self.odt_path)
					request = client.post("/upload", auth=self.auth, data={"file": (file, filename)}, content_type="multipart/form-data")
				self.assertEqual(request.status_code, 400)
				self.assertEqual(json.loads(request.data)["message"], "Only Excel files (.xlsx) are allowed")
	def test_file_empty(self):
		with self.client as client:
			with self.app_context():
				with open(self.excel_empty_path, "rb") as file:
					filename = os.path.basename(self.excel_empty_path)
					request = client.post("/upload", auth=self.auth, data={"file": (file, filename)}, content_type="multipart/form-data")
				self.assertEqual(request.status_code, 400)
				self.assertEqual(json.loads(request.data)["message"], "Uploaded file is not a valid Excel file. Please check file is not empty")
	def test_file_wrong_format(self):
		with self.client as client:
			with self.app_context():
				with open(self.excel_wrong_format_path, "rb") as file:
					filename = os.path.basename(self.excel_wrong_format_path)
					request = client.post("/upload", auth=self.auth, data={"file": (file, filename)}, content_type="multipart/form-data")
				self.assertEqual(request.status_code, 400)
				self.assertEqual(json.loads(request.data)["message"], "Excel format is not correct.")

	def tearDown(self):
		super().tearDown() #calling parent class setUp to avoid overwriting it with the instructions below
		os.remove(self.excel_path)
		os.remove(self.excel_wrong_format_path)
		os.remove(self.odt_path)
		os.remove(self.excel_empty_path)
		self.assertFalse(os.path.exists(self.excel_path))
		self.assertFalse(os.path.exists(self.odt_path))
		self.assertFalse(os.path.exists(self.excel_wrong_format_path))
		self.assertFalse(os.path.exists(self.excel_empty_path))